import click
import MySQLdb
from MySQLdb import Error
from openpyxl import load_workbook
import django
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE','classproject.settings')
django.setup()
from onlineapp.models import *

# try:
#     connection = MySQLdb.connect("localhost","user1","Nagasai.41")
#     print("connected to Database")
# except Error as e:
#     print("can't connect to database,\nError:",e)
#     exit()


@click.group(chain=True)
def cli():
    pass

def close(conn,err):
    click.echo("Error %s" % err)
    conn.close()
    exit()

@cli.command('createdb')
@click.argument('dbname', nargs=1)
def createdb(dbname):
    try:
        cursor = connection.cursor()
        cursor.execute("create database "+dbname)
        cursor.execute("use "+dbname)
        cursor.execute("create table students(name varchar(50),college varchar(10),emailid varchar(60),dbnames varchar(20) PRIMARY KEY)")
        cursor.execute("create table marks(name varchar(20),prob1 int,prob2 int ,prob3 int ,prob4 int,total int,FOREIGN KEY fk(name)REFERENCES students(dbnames) ON UPDATE CASCADE )")
        connection.commit()
        click.echo("Database and tables are created successful")
    except Error as err:
        close(connection,err)
    connection.close()


@cli.command('dropdb')
@click.argument('dbname', nargs=1)
def dropdb(dbname):
    try:
        connection.cursor().execute("drop database "+dbname)
        click.echo("database dropped Success")
    except Error as err:
        close(connection,err)
    connection.close()


@cli.command('importdata')
@click.argument('dbname', nargs=1)
@click.argument('stdfile', nargs=1)
@click.argument('marksfile', nargs=1)
def importdata(dbname, stdfile, marksfile):


    wb = load_workbook(stdfile)
    for row in wb["Colleges"].values:
        if row[3]=="Contact":
            continue
        ins=College(name=row[0],acronym=row[1],location=row[2],contact=row[3])
        ins.save()
        print(row[0])
    for row in wb["Current"].values:
        if row[0] == "Name":
            continue
        #query = "insert into students values('"+row[0]+"','"+row[1]+"','"+row[2]+"','"+row[3].lower()+"')"
        ins=Student(name=row[0],college=College.objects.get(acronym=row[1]),email=row[2],db_folder=str.lower(row[3]));
        ins.save();
        # try:
        #     cur.execute(query)
        # except Error as err:
        #     close(connection, err)
    wb = load_workbook(marksfile)
    for row in wb["Sheet"].values:
        if row[0] == " Student " or ''.join(row[0].split('_')[2]) == 'skc':
            continue
        ins=MockTest1(student=Student.objects.get(db_folder=row[0].split('_')[2]),problem1=row[1],problem2=row[2],problem3=row[3],problem4=row[4],total=row[5])
        ins.save()
        # query = "insert into marks values('"+''.join(row[0].split('_')[2])+"',"+row[1]+","+row[2]+","+row[3]+","+row[4]+","+row[5]+")"
        # try:
        #     cur.execute(query)
        # except Error as err:
        #     close(connection, err)
    #connection.commit()
    click.echo("Successfully Imported")


@cli.command('clgstats')
@click.argument('dbname', nargs=1)
def clgstats(dbname):
    # try:
    #     # cur = connection.cursor()
    #     # cur.execute("use "+dbname)
    #     # cur.execute("select s.college,count(*),min(m.total),avg(m.total),max(m.total) from students s,marks m where s.dbnames=m.name group by s.college")
    #     # click.echo("College\t   StdCount  Min.Marks Avg.Marks\tMax.Marks")
    #     # for row in cur.fetchall():
    #     #     click.echo("{0}\t\t{1}\t{2}\t{3}\t\t{4}" .format(row[0], row[1], row[2], row[3], row[4]))
    # except Error as err:
    # #     close(connection, err)
    # # connection.close()
    pass

cli()
